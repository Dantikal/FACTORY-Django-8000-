from html import escape
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import generics

from .models import Invoice
from .serializers import InvoiceSerializer
from shared.permissions import IsFactory


class InvoiceListView(generics.ListAPIView):
    queryset = Invoice.objects.prefetch_related('items').all().order_by('-created_at')
    serializer_class = InvoiceSerializer
    permission_classes = [IsFactory]


class InvoiceDetailView(generics.RetrieveAPIView):
    queryset = Invoice.objects.prefetch_related('items').all()
    serializer_class = InvoiceSerializer
    permission_classes = [IsFactory]
    lookup_url_kwarg = 'id'


class InvoicePdfView(generics.RetrieveAPIView):
    queryset = Invoice.objects.prefetch_related('items').all()
    permission_classes = [IsFactory]
    lookup_url_kwarg = 'id'

    def get(self, request, *args, **kwargs):
        invoice = self.get_object()
        pdf = self._render_pdf(invoice)
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice-{invoice.id}.pdf"'
        return response

    def _render_pdf(self, invoice):
        try:
            from weasyprint import HTML
            return HTML(string=self._build_html(invoice)).write_pdf()
        except (ImportError, OSError):
            return self._build_fallback_pdf(invoice)

    def _build_html(self, invoice):
        created_at = timezone.localtime(invoice.created_at).strftime('%Y-%m-%d %H:%M')
        rows = ''.join(
            (
                '<tr>'
                f'<td>{escape(str(item.product_id))}</td>'
                f'<td class="num">{item.qty}</td>'
                f'<td class="num">{item.price}</td>'
                f'<td class="num">{item.total}</td>'
                '</tr>'
            )
            for item in invoice.items.all()
        )
        return f"""
        <!doctype html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: sans-serif; font-size: 12px; color: #111827; }}
                h1 {{ font-size: 22px; margin: 0 0 16px; }}
                .meta {{ margin-bottom: 18px; }}
                .meta div {{ margin: 4px 0; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ border: 1px solid #d1d5db; padding: 7px; text-align: left; }}
                th {{ background: #f3f4f6; }}
                .num {{ text-align: right; }}
                .total {{ margin-top: 16px; text-align: right; font-weight: 700; }}
            </style>
        </head>
        <body>
            <h1>Invoice {escape(str(invoice.id))}</h1>
            <div class="meta">
                <div><strong>Dispatch ID:</strong> {escape(str(invoice.dispatch_id))}</div>
                <div><strong>Driver ID:</strong> {escape(str(invoice.driver_id))}</div>
                <div><strong>Warehouse ID:</strong> {escape(str(invoice.warehouse_id))}</div>
                <div><strong>Created at:</strong> {created_at}</div>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Product ID</th>
                        <th>Qty</th>
                        <th>Price</th>
                        <th>Total</th>
                    </tr>
                </thead>
                <tbody>{rows}</tbody>
            </table>
            <div class="total">Total amount: {invoice.total_amount}</div>
        </body>
        </html>
        """

    def _build_fallback_pdf(self, invoice):
        created_at = timezone.localtime(invoice.created_at).strftime('%Y-%m-%d %H:%M')
        lines = [
            f'Invoice {invoice.id}',
            f'Dispatch ID: {invoice.dispatch_id}',
            f'Driver ID: {invoice.driver_id}',
            f'Warehouse ID: {invoice.warehouse_id}',
            f'Created at: {created_at}',
            '',
            'Product ID                              Qty        Price        Total',
        ]
        for item in invoice.items.all():
            lines.append(
                f'{str(item.product_id)[:36]:36} {item.qty:>8} {str(item.price):>12} {str(item.total):>12}'
            )
        lines.extend(['', f'Total amount: {invoice.total_amount}'])

        content = ['BT', '/F1 10 Tf', '50 800 Td', '12 TL']
        for line in lines[:60]:
            content.append(f'({self._escape_pdf_text(line[:95])}) Tj')
            content.append('T*')
        content.append('ET')
        stream = '\n'.join(content).encode('latin-1', 'replace')

        objects = [
            b'<< /Type /Catalog /Pages 2 0 R >>',
            b'<< /Type /Pages /Kids [3 0 R] /Count 1 >>',
            b'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>',
            b'<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>',
            b'<< /Length ' + str(len(stream)).encode('ascii') + b' >>\nstream\n' + stream + b'\nendstream',
        ]

        pdf = b'%PDF-1.4\n'
        offsets = [0]
        for index, obj in enumerate(objects, start=1):
            offsets.append(len(pdf))
            pdf += f'{index} 0 obj\n'.encode('ascii') + obj + b'\nendobj\n'
        xref_offset = len(pdf)
        pdf += f'xref\n0 {len(objects) + 1}\n'.encode('ascii')
        pdf += b'0000000000 65535 f \n'
        for offset in offsets[1:]:
            pdf += f'{offset:010d} 00000 n \n'.encode('ascii')
        pdf += (
            b'trailer\n'
            + f'<< /Size {len(objects) + 1} /Root 1 0 R >>\n'.encode('ascii')
            + b'startxref\n'
            + f'{xref_offset}\n'.encode('ascii')
            + b'%%EOF\n'
        )
        return pdf

    def _escape_pdf_text(self, value):
        return value.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


class InvoiceExcelView(generics.RetrieveAPIView):
    queryset = Invoice.objects.prefetch_related('items').all()
    permission_classes = [IsFactory]
    lookup_url_kwarg = 'id'

    def get(self, request, *args, **kwargs):
        invoice = self.get_object()
        workbook = self._build_workbook(invoice)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="invoice-{invoice.id}.xlsx"'
        return response

    def _build_workbook(self, invoice):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Invoice'

        sheet['A1'] = 'Invoice'
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A3'] = 'ID'
        sheet['B3'] = str(invoice.id)
        sheet['A4'] = 'Dispatch ID'
        sheet['B4'] = str(invoice.dispatch_id)
        sheet['A5'] = 'Driver ID'
        sheet['B5'] = str(invoice.driver_id)
        sheet['A6'] = 'Warehouse ID'
        sheet['B6'] = str(invoice.warehouse_id)
        sheet['A7'] = 'Created at'
        sheet['B7'] = timezone.localtime(invoice.created_at).strftime('%Y-%m-%d %H:%M')

        headers = ['Product ID', 'Qty', 'Price', 'Total']
        header_row = 9
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=column, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='E5E7EB')

        for row, item in enumerate(invoice.items.all(), start=header_row + 1):
            sheet.cell(row=row, column=1, value=str(item.product_id))
            sheet.cell(row=row, column=2, value=item.qty)
            sheet.cell(row=row, column=3, value=float(item.price))
            sheet.cell(row=row, column=4, value=float(item.total))

        total_row = header_row + invoice.items.count() + 2
        sheet.cell(row=total_row, column=3, value='Total amount').font = Font(bold=True)
        sheet.cell(row=total_row, column=4, value=float(invoice.total_amount)).font = Font(bold=True)

        widths = [38, 12, 14, 14]
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width

        return workbook
