from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.db.models import Count, DecimalField, ExpressionWrapper, F, IntegerField, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView

from apps.invoices.models import Invoice
from apps.payments.models import FactoryPayment, WarehouseDebt
from apps.products.models import Product
from apps.reception.models import DeliveryItem, FactoryDelivery
from apps.shipments.models import Shipment
from .serializers import ReportQuerySerializer
from shared.permissions import IsManager, IsAccountant, IsAdmin


DECIMAL_ZERO = Decimal('0.00')


def _as_text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _decimal_sum(queryset, field):
    return queryset.aggregate(total=Sum(field))['total'] or DECIMAL_ZERO


def _apply_date_filter(queryset, field, filters, is_datetime=False):
    lookup = f'{field}__date' if is_datetime else field
    date_from = filters.get('date_from')
    date_to = filters.get('date_to')

    if date_from:
        queryset = queryset.filter(**{f'{lookup}__gte': date_from})
    if date_to:
        queryset = queryset.filter(**{f'{lookup}__lte': date_to})

    return queryset


def _apply_warehouse_filter(queryset, filters):
    warehouse_id = filters.get('warehouse_id')
    if warehouse_id:
        return queryset.filter(warehouse_id=warehouse_id)
    return queryset


def _quantity_expression(prefix=''):
    return Coalesce(f'{prefix}actual_qty', f'{prefix}expected_qty', output_field=IntegerField())


def _amount_expression(quantity, price_field):
    return ExpressionWrapper(
        quantity * F(price_field),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )


class BaseXlsxReportView(APIView):
    permission_classes = [IsManager | IsAccountant | IsAdmin]

    report_title = 'Report'
    filename = 'report.xlsx'

    def get(self, request):
        query_serializer = ReportQuerySerializer(data=request.query_params)
        query_serializer.is_valid(raise_exception=True)
        filters = query_serializer.validated_data
        workbook = self.build_workbook(filters)
        return self.xlsx_response(workbook)

    def build_workbook(self, filters):
        raise NotImplementedError

    def workbook(self, title, headers, rows, filters, notes=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title[:31]

        sheet['A1'] = title
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A2'] = 'Generated at'
        sheet['B2'] = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
        sheet['A3'] = 'date_from'
        sheet['B3'] = _as_text(filters.get('date_from'))
        sheet['A4'] = 'date_to'
        sheet['B4'] = _as_text(filters.get('date_to'))
        sheet['A5'] = 'warehouse_id'
        sheet['B5'] = _as_text(filters.get('warehouse_id'))

        current_row = 7
        if notes:
            for note in notes:
                sheet.cell(row=current_row, column=1, value=note)
                current_row += 1
            current_row += 1

        header_row = current_row
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=column, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='E5E7EB')

        data_start = header_row + 1
        rows = list(rows)
        if rows:
            for row_number, row in enumerate(rows, start=data_start):
                for column, value in enumerate(row, start=1):
                    sheet.cell(row=row_number, column=column, value=_as_text(value))
        else:
            sheet.cell(row=data_start, column=1, value='No data')

        sheet.freeze_panes = sheet.cell(row=data_start, column=1)
        for column in range(1, len(headers) + 1):
            letter = get_column_letter(column)
            width = max(
                len(str(sheet.cell(row=row, column=column).value or ''))
                for row in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[letter].width = min(max(width + 2, 12), 42)

        return workbook

    def xlsx_response(self, workbook):
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        return response

    def external_json(self, url, params):
        try:
            import requests
            response = requests.get(url, params=params, timeout=5)
            response.raise_for_status()
            data = response.json()
            if isinstance(data, list):
                return data
            if isinstance(data, dict):
                return data.get('results', [])
            return []
        except Exception:
            return None


class ShipmentReportView(BaseXlsxReportView):
    report_title = 'Shipments'
    filename = 'shipments-report.xlsx'

    def build_workbook(self, filters):
        queryset = _apply_warehouse_filter(Shipment.objects.all().order_by('-shipment_date'), filters)
        queryset = _apply_date_filter(queryset, 'shipment_date', filters)
        headers = [
            'id', 'warehouse_id', 'shipment_date', 'truck_number', 'truck_driver',
            'status', 'total_amount', 'created_by_id', 'created_at', 'updated_at',
        ]
        rows = (
            [
                item.id, item.warehouse_id, item.shipment_date, item.truck_number,
                item.truck_driver, item.status, item.total_amount, item.created_by_id,
                item.created_at, item.updated_at,
            ]
            for item in queryset
        )
        return self.workbook(self.report_title, headers, rows, filters)


class ReceptionReportView(BaseXlsxReportView):
    report_title = 'Receptions'
    filename = 'receptions-report.xlsx'

    def build_workbook(self, filters):
        queryset = _apply_warehouse_filter(FactoryDelivery.objects.all().order_by('-delivered_at'), filters)
        queryset = _apply_date_filter(queryset, 'delivered_at', filters, is_datetime=True)
        headers = [
            'id', 'shipment_id', 'warehouse_id', 'delivery_number', 'delivered_at',
            'status', 'total_amount', 'warehouse_comment', 'created_by_id',
            'client_id', 'operation_time', 'created_at', 'updated_at',
        ]
        rows = (
            [
                item.id, item.shipment_id, item.warehouse_id, item.delivery_number,
                item.delivered_at, item.status, item.total_amount, item.warehouse_comment,
                item.created_by_id, item.client_id, item.operation_time,
                item.created_at, item.updated_at,
            ]
            for item in queryset
        )
        return self.workbook(self.report_title, headers, rows, filters)


class InventoryReportView(BaseXlsxReportView):
    report_title = 'Inventory'
    filename = 'inventory-report.xlsx'

    def build_workbook(self, filters):
        params = {}
        if filters.get('warehouse_id'):
            params['warehouse_id'] = str(filters['warehouse_id'])

        inventory = self.external_json(f'{settings.WAREHOUSE_SERVICE_URL}/api/warehouse/inventory', params)
        headers = [
            'product_id', 'barcode', 'name', 'pieces_per_box', 'factory_price',
            'dispatch_price', 'stock_quantity', 'batch_number', 'expiry_date', 'status',
        ]

        products = Product.objects.all().order_by('name')
        inventory_by_barcode = {}
        if inventory is not None:
            inventory_by_barcode = {str(item.get('barcode')): item for item in inventory}

        rows = (
            [
                product.id,
                product.barcode,
                product.name,
                getattr(product, 'pieces_per_box', ''),
                product.factory_price,
                product.dispatch_price,
                inventory_by_barcode.get(product.barcode, {}).get('stockQuantity', 'not_available'),
                product.batch_number,
                product.expiry_date,
                product.status,
            ]
            for product in products
        )
        notes = None if inventory is not None else ['warehouse inventory service is unavailable']
        return self.workbook(self.report_title, headers, rows, filters, notes=notes)


class RegionalSalesReportView(BaseXlsxReportView):
    report_title = 'Regional Sales'
    filename = 'regional-sales-report.xlsx'

    def build_workbook(self, filters):
        queryset = _apply_warehouse_filter(FactoryDelivery.objects.all(), filters)
        queryset = _apply_date_filter(queryset, 'delivered_at', filters, is_datetime=True)
        rows_data = (
            queryset
            .values('warehouse_id')
            .annotate(deliveries=Count('id'), total_sales=Coalesce(Sum('total_amount'), Value(DECIMAL_ZERO)))
            .order_by('-total_sales')
        )
        headers = ['warehouse_id', 'deliveries', 'total_sales']
        rows = ([row['warehouse_id'], row['deliveries'], row['total_sales']] for row in rows_data)
        return self.workbook(self.report_title, headers, rows, filters)


class WarehouseDebtReportView(BaseXlsxReportView):
    report_title = 'Warehouse Debts'
    filename = 'warehouse-debts-report.xlsx'

    def build_workbook(self, filters):
        queryset = _apply_warehouse_filter(WarehouseDebt.objects.all().order_by('-amount'), filters)
        queryset = _apply_date_filter(queryset, 'updated_at', filters, is_datetime=True)
        headers = ['id', 'warehouse_id', 'amount', 'created_at', 'updated_at']
        rows = ([item.id, item.warehouse_id, item.amount, item.created_at, item.updated_at] for item in queryset)
        return self.workbook(self.report_title, headers, rows, filters)


class ProductRatingReportView(BaseXlsxReportView):
    report_title = 'Product Rating'
    filename = 'product-rating-report.xlsx'

    def build_workbook(self, filters):
        queryset = DeliveryItem.objects.select_related('product', 'delivery')
        if filters.get('warehouse_id'):
            queryset = queryset.filter(delivery__warehouse_id=filters['warehouse_id'])
        queryset = _apply_date_filter(queryset, 'delivery__delivered_at', filters, is_datetime=True)

        quantity = _quantity_expression()
        total_amount = _amount_expression(quantity, 'product__dispatch_price')
        rows_data = (
            queryset
            .values('product_id', 'product__barcode', 'product__name')
            .annotate(
                quantity=Coalesce(Sum(quantity), Value(0), output_field=IntegerField()),
                total_amount=Coalesce(
                    Sum(total_amount),
                    Value(DECIMAL_ZERO),
                    output_field=DecimalField(max_digits=14, decimal_places=2),
                ),
            )
            .order_by('-quantity', '-total_amount')
        )
        headers = ['product_id', 'barcode', 'name', 'quantity', 'total_amount']
        rows = (
            [
                row['product_id'], row['product__barcode'], row['product__name'],
                row['quantity'], row['total_amount'],
            ]
            for row in rows_data
        )
        return self.workbook(self.report_title, headers, rows, filters)


class MovementReportView(BaseXlsxReportView):
    report_title = 'Movements'
    filename = 'movements-report.xlsx'

    def build_workbook(self, filters):
        shipments = _apply_warehouse_filter(Shipment.objects.all(), filters)
        shipments = _apply_date_filter(shipments, 'shipment_date', filters)
        receptions = _apply_warehouse_filter(FactoryDelivery.objects.all(), filters)
        receptions = _apply_date_filter(receptions, 'delivered_at', filters, is_datetime=True)

        rows = []
        for shipment in shipments:
            rows.append([
                'shipment', shipment.id, shipment.shipment_date, shipment.warehouse_id,
                shipment.status, shipment.total_amount,
            ])
        for reception in receptions:
            rows.append([
                'reception', reception.id, reception.delivered_at, reception.warehouse_id,
                reception.status, reception.total_amount,
            ])

        rows.sort(key=lambda row: _as_text(row[2]), reverse=True)
        headers = ['type', 'id', 'date', 'warehouse_id', 'status', 'total_amount']
        return self.workbook(self.report_title, headers, rows, filters)


class DriverDebtReportView(BaseXlsxReportView):
    report_title = 'Driver Debts'
    filename = 'driver-debts-report.xlsx'

    def build_workbook(self, filters):
        params = {}
        if filters.get('warehouse_id'):
            params['warehouse_id'] = str(filters['warehouse_id'])

        data = self.external_json(f'{settings.DRIVERS_SERVICE_URL}/api/drivers/debts', params)
        headers = ['driver_id', 'full_name', 'car_number', 'warehouse_id', 'total_debt', 'updated_at']
        rows = []
        if data is not None:
            rows = [
                [
                    item.get('driverId') or item.get('driver_id'),
                    item.get('fullName') or item.get('full_name'),
                    item.get('carNumber') or item.get('car_number'),
                    item.get('warehouseId') or item.get('warehouse_id'),
                    item.get('totalDebt') or item.get('total_debt') or item.get('debt'),
                    item.get('updatedAt') or item.get('updated_at'),
                ]
                for item in data
            ]
        notes = None if data is not None else ['drivers service is unavailable']
        return self.workbook(self.report_title, headers, rows, filters, notes=notes)


class CashboxReportView(BaseXlsxReportView):
    report_title = 'Cashbox'
    filename = 'cashbox-report.xlsx'

    def build_workbook(self, filters):
        queryset = _apply_warehouse_filter(FactoryPayment.objects.all().order_by('-paid_at'), filters)
        queryset = _apply_date_filter(queryset, 'paid_at', filters, is_datetime=True)
        headers = [
            'id', 'warehouse_id', 'amount', 'payment_method', 'comment',
            'paid_at', 'created_by_id', 'client_id', 'operation_time', 'created_at', 'updated_at',
        ]
        rows = (
            [
                item.id, item.warehouse_id, item.amount, item.payment_method, item.comment,
                item.paid_at, item.created_by_id, item.client_id, item.operation_time,
                item.created_at, item.updated_at,
            ]
            for item in queryset
        )
        return self.workbook(self.report_title, headers, rows, filters)


class DispatchReportView(BaseXlsxReportView):
    report_title = 'Dispatches'
    filename = 'dispatches-report.xlsx'

    def build_workbook(self, filters):
        queryset = _apply_warehouse_filter(Invoice.objects.all().order_by('-created_at'), filters)
        queryset = _apply_date_filter(queryset, 'created_at', filters, is_datetime=True)
        headers = ['id', 'dispatch_id', 'driver_id', 'warehouse_id', 'total_amount', 'created_at', 'updated_at']
        rows = (
            [
                item.id, item.dispatch_id, item.driver_id, item.warehouse_id,
                item.total_amount, item.created_at, item.updated_at,
            ]
            for item in queryset
        )
        return self.workbook(self.report_title, headers, rows, filters)


class ReturnReportView(BaseXlsxReportView):
    report_title = 'Returns'
    filename = 'returns-report.xlsx'

    def build_workbook(self, filters):
        params = {}
        for key in ('date_from', 'date_to', 'warehouse_id'):
            if filters.get(key):
                params[key] = str(filters[key])

        data = self.external_json(f'{settings.DRIVERS_SERVICE_URL}/api/returns', params)
        headers = ['id', 'driver_id', 'driver_name', 'warehouse_id', 'total_amount', 'status', 'returned_at', 'created_at']
        rows = []
        if data is not None:
            rows = [
                [
                    item.get('id'),
                    item.get('driverId') or item.get('driver_id'),
                    item.get('driverName') or item.get('driver_name'),
                    item.get('warehouseId') or item.get('warehouse_id'),
                    item.get('totalAmount') or item.get('total_amount'),
                    item.get('status'),
                    item.get('returnedAt') or item.get('returned_at'),
                    item.get('createdAt') or item.get('created_at'),
                ]
                for item in data
            ]
        notes = None if data is not None else ['returns source is unavailable in factory-service']
        return self.workbook(self.report_title, headers, rows, filters, notes=notes)


class LowStockReportView(BaseXlsxReportView):
    report_title = 'Low Stock'
    filename = 'low-stock-report.xlsx'
    default_threshold = 10

    def build_workbook(self, filters):
        params = {}
        if filters.get('warehouse_id'):
            params['warehouse_id'] = str(filters['warehouse_id'])

        inventory = self.external_json(f'{settings.WAREHOUSE_SERVICE_URL}/api/warehouse/inventory', params)
        headers = ['barcode', 'name', 'warehouse_id', 'stock_quantity', 'threshold']
        rows = []
        if inventory is not None:
            product_names = dict(Product.objects.values_list('barcode', 'name'))
            for item in inventory:
                raw_stock_quantity = item.get('stockQuantity') or item.get('stock_quantity') or 0
                try:
                    stock_quantity = int(raw_stock_quantity)
                except (TypeError, ValueError):
                    stock_quantity = 0
                if stock_quantity <= self.default_threshold:
                    barcode = item.get('barcode')
                    rows.append([
                        barcode,
                        item.get('name') or product_names.get(str(barcode), ''),
                        item.get('warehouseId') or item.get('warehouse_id') or filters.get('warehouse_id'),
                        stock_quantity,
                        self.default_threshold,
                    ])

        notes = None if inventory is not None else ['warehouse inventory service is unavailable']
        return self.workbook(self.report_title, headers, rows, filters, notes=notes)
